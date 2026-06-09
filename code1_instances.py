#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jun  9 18:49:53 2026

@author: melisaozdemir
"""

#!/usr/bin/env python3
"""
Kod 1 — ARP-SA Instance Üreticisi
20 test instance üretir.
Çıktı:
  · instances/  → JSON dosyaları (Kod 2 için girdi)
  · ARP_SA_Instances.xlsx → insan okunabilir Excel
"""
import os, json, random, math
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

# ── Paylaşılan altyapı ────────────────────────────────────────
TAU   = ['M','E','W']
OPT   = {'M':['WM','sM1','sM2'],'E':['WE','sE1','sE2'],'W':['WW','sW1','sW2']}
INT   = {'M':'WM','E':'WE','W':'WW'}
NODES = ['depot','sM1','sM2','sE1','sE2','sW1','sW2']
NI    = {n:i for i,n in enumerate(NODES)}
_TT   = [[0.0,1.3,2.1,0.9,1.7,1.5,2.4],[1.3,0.0,1.4,1.8,2.3,2.1,2.9],
          [2.1,1.4,0.0,2.5,1.9,2.7,1.6],[0.9,1.8,2.5,0.0,1.1,1.3,2.0],
          [1.7,2.3,1.9,1.1,0.0,1.8,1.4],[1.5,2.1,2.7,1.3,1.8,0.0,1.2],
          [2.4,2.9,1.6,2.0,1.4,1.2,0.0]]
def tt(i,j): return _TT[NI[i]][NI[j]]
BASE_SU = {'WM':1.5,'sM1':1.2,'sM2':1.8,'WE':0.8,'sE1':0.6,'sE2':1.0,'WW':1.0,'sW1':0.9,'sW2':1.1}
FIX_T   = ['F1','F2','F3','F4']
PROC_D  = {'M':{'WM':(6.8,1.6),'sM1':(5.5,1.3),'sM2':(8.3,1.8)},
            'E':{'WE':(5.0,1.1),'sE1':(4.1,0.9),'sE2':(5.9,1.2)},
            'W':{'WW':(4.0,0.9),'sW1':(4.6,1.0),'sW2':(3.7,0.8)}}

# ── M01: elle tasarlanmış referans instance ───────────────────
M01 = {
    'name':'M01','n_orders':10,'n_vehicles':2,'difficulty':'Kolay (Ref.)',
    'orders':['P01','P02','P03','P04','P05','P06','P07','P08','P09','P10'],
    'due':{'P01':12,'P02':24,'P03':32,'P04':8,'P05':48,'P06':25,'P07':20,'P08':28,'P09':22,'P10':34},
    'proc':{
        'P01':{'M':{'WM':6.5,'sM1':5.2,'sM2':7.8},'E':{'WE':3.8,'sE1':3.1,'sE2':5.4},'W':{'WW':3.2,'sW1':3.9,'sW2':2.7}},
        'P02':{'M':{'WM':8.1,'sM1':6.4,'sM2':9.3},'E':{'WE':5.6,'sE1':4.7,'sE2':6.2},'W':{'WW':4.1,'sW1':3.6,'sW2':4.8}},
        'P03':{'M':{'WM':5.9,'sM1':4.8,'sM2':6.5},'E':{'WE':7.2,'sE1':6.1,'sE2':8.0},'W':{'WW':5.5,'sW1':5.9,'sW2':4.8}},
        'P04':{'M':{'WM':4.3,'sM1':3.7,'sM2':5.1},'E':{'WE':3.2,'sE1':2.8,'sE2':3.9},'W':{'WW':2.8,'sW1':2.4,'sW2':3.3}},
        'P05':{'M':{'WM':9.7,'sM1':8.2,'sM2':11.1},'E':{'WE':6.4,'sE1':5.5,'sE2':7.3},'W':{'WW':6.8,'sW1':7.2,'sW2':6.1}},
        'P06':{'M':{'WM':7.2,'sM1':5.9,'sM2':8.4},'E':{'WE':4.9,'sE1':4.2,'sE2':5.7},'W':{'WW':4.4,'sW1':4.8,'sW2':3.9}},
        'P07':{'M':{'WM':5.1,'sM1':4.3,'sM2':5.8},'E':{'WE':4.1,'sE1':3.5,'sE2':4.7},'W':{'WW':3.6,'sW1':3.2,'sW2':4.1}},
        'P08':{'M':{'WM':7.8,'sM1':6.5,'sM2':8.9},'E':{'WE':5.3,'sE1':4.6,'sE2':6.1},'W':{'WW':5.1,'sW1':5.5,'sW2':4.6}},
        'P09':{'M':{'WM':6.2,'sM1':5.1,'sM2':7.0},'E':{'WE':4.7,'sE1':3.9,'sE2':5.4},'W':{'WW':3.8,'sW1':4.2,'sW2':3.4}},
        'P10':{'M':{'WM':8.5,'sM1':7.1,'sM2':9.6},'E':{'WE':6.8,'sE1':5.7,'sE2':7.5},'W':{'WW':5.9,'sW1':6.3,'sW2':5.2}},
    },
    'fixtures':{'P01':{'M':'F2','E':'F1','W':'F3'},'P02':{'M':'F1','E':'F3','W':'F2'},
                'P03':{'M':'F3','E':'F2','W':'F1'},'P04':{'M':'F2','E':'F4','W':'F3'},
                'P05':{'M':'F4','E':'F1','W':'F2'},'P06':{'M':'F1','E':'F2','W':'F4'},
                'P07':{'M':'F3','E':'F4','W':'F1'},'P08':{'M':'F2','E':'F3','W':'F2'},
                'P09':{'M':'F4','E':'F1','W':'F3'},'P10':{'M':'F1','E':'F2','W':'F4'}},
    'setup':dict(BASE_SU),'seed':0,'tightness':'custom','complexity':1.0,
}

# ── 20 instance tanımı ────────────────────────────────────────
# (name, n, seed, tightness, complexity, difficulty_label)
SPECS = [
    # Küçük — n=5
    ('S01', 5,101,1.90,0.80,'Kolay'),
    ('S02', 5,102,1.45,0.85,'Kolay-Orta'),
    ('S03', 5,103,1.10,0.90,'Orta'),
    ('S04', 5,104,0.82,1.00,'Zor'),
    # Küçük-Orta — n=7
    ('S05', 7,105,1.80,0.85,'Kolay'),
    ('S06', 7,106,1.35,0.90,'Orta'),
    ('S07', 7,107,1.00,0.95,'Orta-Zor'),
    ('S08', 7,108,0.78,1.05,'Zor'),
    # Orta — n=10 (M01 ref + 4 üretilmiş)
    # M01 ayrı tanımlandı
    ('M02',10,201,1.50,0.95,'Kolay'),
    ('M03',10,202,1.18,1.00,'Orta'),
    ('M04',10,203,0.90,1.05,'Orta-Zor'),
    ('M05',10,204,0.75,1.10,'Zor'),
    # Orta-Büyük — n=12
    ('M06',12,205,1.60,0.95,'Kolay'),
    ('M07',12,206,1.20,1.00,'Orta'),
    ('M08',12,207,0.85,1.10,'Zor'),
    # Büyük — n=15
    ('L01',15,301,1.90,0.90,'Kolay'),
    ('L02',15,302,1.35,1.00,'Orta'),
    ('L03',15,303,0.95,1.05,'Orta-Zor'),
    ('L04',15,304,0.80,1.15,'Zor'),
    # Çok Büyük — n=20
    ('L05',20,305,1.50,0.95,'Kolay-Orta'),
    ('L06',20,306,0.88,1.10,'Zor'),
]

def _clip(v,lo,hi): return max(lo,min(hi,v))

def generate(name,n,seed,tight,cmplx,diff_label):
    rng=random.Random(seed)
    orders=[f'P{i+1:02d}' for i in range(n)]
    proc={}
    for p in orders:
        proc[p]={}
        for tau in TAU:
            proc[p][tau]={}
            for s in OPT[tau]:
                mu,sg=PROC_D[tau][s]
                proc[p][tau][s]=round(_clip(rng.gauss(mu*cmplx,sg*0.6),1.5*cmplx,14.*cmplx),1)
    fixtures={p:{tau:rng.choice(FIX_T) for tau in TAU} for p in orders}
    # Teorik minimum C_p
    theo={}
    for p in orders:
        rm={}
        for tau in TAU:
            br=float('inf')
            for s in OPT[tau]:
                fin=proc[p][tau][s]
                if s!=INT[tau]:
                    arr=tt('depot',s); d=max(arr,fin); r=d+tt(s,'depot')
                else: r=fin
                br=min(br,r)
            rm[tau]=br
        theo[p]=max(rm.values())
    # Kuyruk yükü tahmini
    q_wait=max(0.,(n/3-1)/2)*6.5*cmplx*0.55
    due={}
    for p in orders:
        base=(theo[p]+q_wait)*tight
        due[p]=max(5,round(base+rng.uniform(-.08,.10)*base))
    return {'name':name,'n_orders':n,'n_vehicles':2,'difficulty':diff_label,
            'orders':orders,'due':due,'proc':proc,'fixtures':fixtures,
            'setup':dict(BASE_SU),'seed':seed,'tightness':tight,'complexity':cmplx}

def build_all():
    out={'M01':M01}
    for nm,n,sd,ti,cm,di in SPECS:
        out[nm]=generate(nm,n,sd,ti,cm,di)
    return out

# ── Excel renk/stil yardımcıları ─────────────────────────────
def hdr_fill(hex_): return PatternFill('solid',start_color=hex_)
def thin_border():
    t=Side(style='thin',color='BFBFBF')
    return Border(left=t,right=t,top=t,bottom=t)
def set_hdr(cell,txt,fg='FFFFFF',bg='1F3864',sz=10,bold=True,wrap=False):
    cell.value=txt
    cell.font=Font(name='Calibri',bold=bold,color=fg,size=sz)
    cell.fill=hdr_fill(bg)
    cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=wrap)
    cell.border=thin_border()
def set_data(cell,val,fmt=None,bold=False,align='center'):
    cell.value=val
    cell.font=Font(name='Calibri',bold=bold,size=9)
    cell.alignment=Alignment(horizontal=align,vertical='center')
    cell.border=thin_border()
    if fmt: cell.number_format=fmt
def alt_fill(row_idx):
    return hdr_fill('DCE6F1') if row_idx%2==0 else hdr_fill('FFFFFF')
def set_row_fill(ws,row,col_start,col_end,fill):
    for c in range(col_start,col_end+1):
        ws.cell(row=row,column=c).fill=fill
def autofit(ws,extra=2):
    for col in ws.columns:
        mx=0
        for cell in col:
            try:
                if cell.value: mx=max(mx,len(str(cell.value)))
            except: pass
        ws.column_dimensions[get_column_letter(col[0].column)].width=mx+extra

# ── Sheet 1: Instance Listesi ─────────────────────────────────
def sheet_summary(wb,instances):
    ws=wb.create_sheet('Instance_Listesi'); ws.freeze_panes='A2'
    ws.row_dimensions[1].height=22
    hdrs=['Instance','Sipariş Sayısı (|P|)','Zorluk Sınıfı','Sıkılık (γ)',
          'Karmaşıklık (κ)','Min Due (saat)','Maks Due (saat)','Ort Due (saat)',
          'Araç Sayısı','Oluşturma Yöntemi']
    for c,h in enumerate(hdrs,1): set_hdr(ws.cell(1,c),h,sz=9)
    order=['M01']+[s[0] for s in SPECS]
    for ri,nm in enumerate(order,2):
        inst=instances[nm]
        dues=list(inst['due'].values())
        row=[nm,inst['n_orders'],inst['difficulty'],
             inst.get('tightness','custom'),inst.get('complexity',1.0),
             min(dues),max(dues),round(sum(dues)/len(dues),1),
             inst['n_vehicles'],'Elle tasarım' if nm=='M01' else 'Üretici (seed='+str(inst['seed'])+')']
        fill=alt_fill(ri)
        for c,v in enumerate(row,1):
            set_data(ws.cell(ri,c),v)
            ws.cell(ri,c).fill=fill
    autofit(ws)

# ── Sheet 2: Teslim Tarihleri ─────────────────────────────────
def sheet_due(wb,instances):
    ws=wb.create_sheet('Teslim_Tarihleri'); ws.freeze_panes='B2'
    max_n=max(inst['n_orders'] for inst in instances.values())
    order_cols=[f'P{i+1:02d}' for i in range(max_n)]
    set_hdr(ws.cell(1,1),'Instance')
    for c,p in enumerate(order_cols,2): set_hdr(ws.cell(1,c),p,sz=9)
    order=['M01']+[s[0] for s in SPECS]
    for ri,nm in enumerate(order,2):
        inst=instances[nm]; fill=alt_fill(ri)
        set_data(ws.cell(ri,1),nm,bold=True); ws.cell(ri,1).fill=fill
        for c,p in enumerate(order_cols,2):
            v=inst['due'].get(p,'')
            set_data(ws.cell(ri,c),v,'#,##0.0' if v!='' else None)
            ws.cell(ri,c).fill=fill
    autofit(ws)

# ── Sheet 3: İşlem Süreleri ───────────────────────────────────
def sheet_proc(wb,instances):
    ws=wb.create_sheet('Islem_Sureleri'); ws.freeze_panes='C3'
    all_opts=[s for tau in TAU for s in OPT[tau]]
    # Satır 1: grup başlıkları
    set_hdr(ws.cell(1,1),'Instance'); set_hdr(ws.cell(1,2),'Sipariş')
    for gi,(tau,opts) in enumerate([(t,OPT[t]) for t in TAU]):
        c0=3+gi*3
        ws.merge_cells(start_row=1,start_column=c0,end_row=1,end_column=c0+2)
        bg='2E75B6' if tau=='M' else ('70AD47' if tau=='E' else 'C00000')
        set_hdr(ws.cell(1,c0),f'Parça Türü: {tau} — İşlem Süreleri (saat)',bg=bg)
    # Satır 2: opsiyon başlıkları
    set_hdr(ws.cell(2,1),'',bg='F2F2F2',fg='000000')
    set_hdr(ws.cell(2,2),'',bg='F2F2F2',fg='000000')
    for c,s in enumerate(all_opts,3): set_hdr(ws.cell(2,c),s,bg='BDD7EE',fg='000000',sz=9)
    row_idx=3
    for nm in ['M01']+[s[0] for s in SPECS]:
        inst=instances[nm]
        for p in inst['orders']:
            fill=alt_fill(row_idx)
            set_data(ws.cell(row_idx,1),nm,bold=True); ws.cell(row_idx,1).fill=fill
            set_data(ws.cell(row_idx,2),p);            ws.cell(row_idx,2).fill=fill
            for c,s in enumerate(all_opts,3):
                tau=next(t for t in TAU if s in OPT[t])
                v=inst['proc'].get(p,{}).get(tau,{}).get(s,'')
                set_data(ws.cell(row_idx,c),v,'0.0' if v!='' else None)
                ws.cell(row_idx,c).fill=fill
            row_idx+=1
    autofit(ws)

# ── Sheet 4: Aparat Tipleri ───────────────────────────────────
def sheet_fix(wb,instances):
    ws=wb.create_sheet('Aparat_Tipleri'); ws.freeze_panes='C2'
    for c,h in enumerate(['Instance','Sipariş','M Aparatı','E Aparatı','W Aparatı'],1):
        set_hdr(ws.cell(1,c),h)
    row_idx=2
    for nm in ['M01']+[s[0] for s in SPECS]:
        inst=instances[nm]
        for p in inst['orders']:
            fill=alt_fill(row_idx)
            vals=[nm,p]+[inst['fixtures'].get(p,{}).get(tau,'') for tau in TAU]
            for c,v in enumerate(vals,1):
                set_data(ws.cell(row_idx,c),v,bold=(c<=2))
                ws.cell(row_idx,c).fill=fill
            row_idx+=1
    autofit(ws)

# ── Sheet 5: Altyapı (Setup + Seyahat) ───────────────────────
def sheet_infra(wb):
    ws=wb.create_sheet('Altyapi_Paylasilan')
    # Setup süreleri
    set_hdr(ws.cell(1,1),'Setup Süreleri (Aparat Değişimi)',bg='1F3864')
    ws.merge_cells('A1:B1')
    set_hdr(ws.cell(2,1),'Seçenek (s)'); set_hdr(ws.cell(2,2),'Setup Süresi (saat)')
    for ri,(s,v) in enumerate(BASE_SU.items(),3):
        set_data(ws.cell(ri,1),s); set_data(ws.cell(ri,2),v,'0.0')
        ws.cell(ri,1).fill=alt_fill(ri); ws.cell(ri,2).fill=alt_fill(ri)
    # Seyahat matrisi
    r0=len(BASE_SU)+4
    set_hdr(ws.cell(r0,1),'Seyahat Süresi Matrisi t_ij (saat)',bg='1F3864')
    ws.merge_cells(start_row=r0,start_column=1,end_row=r0,end_column=len(NODES)+1)
    set_hdr(ws.cell(r0+1,1),'',bg='F2F2F2',fg='000000')
    for c,n in enumerate(NODES,2): set_hdr(ws.cell(r0+1,c),n,bg='BDD7EE',fg='000000',sz=9)
    for ri,ni in enumerate(NODES,r0+2):
        set_hdr(ws.cell(ri,1),ni,bg='BDD7EE',fg='000000',sz=9)
        for c,nj in enumerate(NODES,2):
            v=_TT[NI[ni]][NI[nj]]
            set_data(ws.cell(ri,c),v,'0.0')
            ws.cell(ri,c).fill=alt_fill(ri)
    autofit(ws)

# ── Ana fonksiyon ─────────────────────────────────────────────
def main():
    base=os.path.dirname(os.path.abspath(__file__))
    inst_dir=os.path.join(base,'instances')
    os.makedirs(inst_dir,exist_ok=True)

    print("Instance'lar üretiliyor...")
    instances=build_all()

    # JSON kaydet
    for nm,inst in instances.items():
        with open(os.path.join(inst_dir,f'{nm}.json'),'w') as f:
            json.dump(inst,f,indent=2)

    # Excel oluştur
    wb=Workbook()
    wb.remove(wb.active)  # varsayılan sayfayı kaldır
    sheet_summary(wb,instances)
    sheet_due(wb,instances)
    sheet_proc(wb,instances)
    sheet_fix(wb,instances)
    sheet_infra(wb)

    xl_path=os.path.join(base,'ARP_SA_Instances.xlsx')
    wb.save(xl_path)

    print(f"  {len(instances)} instance üretildi")
    print(f"  JSON → {inst_dir}/")
    print(f"  Excel → {xl_path}")

    # Özet tablosu
    print(f"\n  {'Instance':<8} {'|P|':>4}  {'Zorluk':<12}  {'Due [min,maks]':>16}")
    print("  "+"-"*48)
    for nm in ['M01']+[s[0] for s in SPECS]:
        inst=instances[nm]
        dues=list(inst['due'].values())
        print(f"  {nm:<8} {inst['n_orders']:>4}  {inst['difficulty']:<12}  [{min(dues):3d},{max(dues):3d}]")

    return instances

if __name__=='__main__':
    main()