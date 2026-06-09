#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jun  9 19:08:42 2026

@author: melisaozdemir
"""

#!/usr/bin/env python3
"""
Kod 3 — ARP-SA Parametre Optimizasyonu (Grid Search)
Optimize edilen parametreler: q, T0, alpha, rho
Test instance'ları: M01 (kolay-referans) + M03 (orta) + L02 (büyük-orta)
Her kombinasyon 5 seed ile çalıştırılır.
Çıktı: ARP_SA_ParamTuning.xlsx
"""
import os, json, time, math, statistics, itertools
import arp_sa_v2 as _m
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE     = os.path.dirname(os.path.abspath(__file__))
INST_DIR = os.path.join(BASE, 'instances')

# ════════════════════════════════════════════════════════════
#  PARAMETRE IZGARASI
# ════════════════════════════════════════════════════════════
GRID = {
    'q':     [2, 3, 4, 6, 9],
    'T0':    [2.0, 4.0, 6.0, 10.0],
    'alpha': [0.95, 0.97, 0.99],
    'rho':   [0.10, 0.20, 0.30],
}
# Sabit tutulan parametreler
N_ITER  = 500   # tüm kombinasyonlarda aynı
D1, D2, D3 = 3.0, 2.0, 1.0
N_SEEDS = 5

# Test instance'ları (zorluk seviyelerini temsil eder)
TEST_INSTANCES = ['M01', 'M03', 'L02']

# ════════════════════════════════════════════════════════════
#  YARDIMCI FONKSIYONLAR
# ════════════════════════════════════════════════════════════
def _patch(inst):
    orig = {k: getattr(_m, k) for k in ['ORDERS','DUE','PROC','FIX','SU','NVEH']}
    _m.ORDERS = inst['orders']
    _m.DUE    = {p: float(d) for p, d in inst['due'].items()}
    _m.PROC   = inst['proc']
    _m.FIX    = inst['fixtures']
    _m.SU     = inst['setup']
    _m.NVEH   = inst['n_vehicles']
    return orig

def _restore(orig):
    for k, v in orig.items(): setattr(_m, k, v)

def run_combo(inst, q, T0, alpha, rho):
    """Tek (parametre, instance) kombinasyonu — N_SEEDS seed ile çalıştır."""
    results = []
    for seed in range(N_SEEDS):
        orig = _patch(inst)
        try:
            t0 = time.time()
            a, s, r, fb, fi, hist, op = _m.alns(
                n_iter=N_ITER, q=q, T0=T0, alpha=alpha, rho=rho,
                d1=D1, d2=D2, d3=D3, seed=seed, verbose=False)
            cpu = time.time() - t0
            _, _, C, T = _m.evaluate(a, s, r)
        finally:
            _restore(orig)
        results.append({
            'seed': seed, 'f_init': round(fi, 3), 'f_best': round(fb, 3),
            'cpu_s': round(cpu, 3),
            'on_time': sum(1 for p in inst['orders'] if T[p] < 1e-6),
        })
    fb_vals = [r['f_best'] for r in results]
    return {
        'f_init':   results[0]['f_init'],
        'mean':     round(statistics.mean(fb_vals), 3),
        'std':      round(statistics.stdev(fb_vals) if len(fb_vals) > 1 else 0., 3),
        'best':     round(min(fb_vals), 3),
        'worst':    round(max(fb_vals), 3),
        'impr_pct': round(100*(results[0]['f_init']-statistics.mean(fb_vals))
                         / max(results[0]['f_init'], 1e-9), 1),
        'mean_cpu': round(statistics.mean(r['cpu_s'] for r in results), 2),
        'mean_ot':  round(statistics.mean(r['on_time'] for r in results), 1),
        'runs':     results,
    }

# ════════════════════════════════════════════════════════════
#  EXCEL STYLE YARDIMCILARI
# ════════════════════════════════════════════════════════════
def _fill(h): return PatternFill('solid', start_color=h)
def _border():
    t = Side(style='thin', color='BFBFBF')
    return Border(left=t, right=t, top=t, bottom=t)

def sh(cell, txt, fg='FFFFFF', bg='1F3864', sz=9, bold=True, wrap=False):
    cell.value = txt
    cell.font = Font(name='Calibri', bold=bold, color=fg, size=sz)
    cell.fill = _fill(bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    cell.border = _border()

def sd(cell, val, fmt=None, bold=False, align='center', color='000000', bg=None):
    cell.value = val
    cell.font = Font(name='Calibri', bold=bold, size=9, color=color)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = _border()
    if fmt:  cell.number_format = fmt
    if bg:   cell.fill = _fill(bg)

def alt(ri): return 'DCE6F1' if ri % 2 == 0 else 'FFFFFF'

def autofit(ws, extra=2):
    for col in ws.columns:
        mx = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = mx + extra

def rank_color(rank, total):
    """En iyi → koyu yeşil, en kötü → koyu kırmızı, arası gradyan."""
    pct = rank / max(total - 1, 1)
    if pct < 0.20:   return 'C6EFCE', '375623'   # yeşil bg / fg
    elif pct < 0.40: return 'FFEB9C', '9C6500'   # sarı
    elif pct < 0.70: return 'FFFFFF', '000000'   # beyaz
    else:            return 'FFC7CE', '9C0006'   # kırmızı

# ════════════════════════════════════════════════════════════
#  SHEET 1: Tüm Kombinasyonlar — Tam Sonuç Tablosu
# ════════════════════════════════════════════════════════════
def sheet_all_combos(wb, all_data, instances):
    """
    Her satır: bir (instance, q, T0, alpha, rho) kombinasyonu.
    """
    ws = wb.create_sheet('Tum_Kombinasyonlar')
    ws.freeze_panes = 'A3'
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 30

    # Satır 1: üst grup başlıkları
    sh(ws.cell(1, 1), 'Parametre Kombinasyonu', bg='1F3864')
    ws.merge_cells('A1:E1')
    sh(ws.cell(1, 6), 'M01 (Kolay, |P|=10)', bg='375623')
    ws.merge_cells('F1:K1')
    sh(ws.cell(1, 12), 'M03 (Orta, |P|=10)', bg='833C00')
    ws.merge_cells('L1:Q1')
    sh(ws.cell(1, 18), 'L02 (Büyük-Orta, |P|=15)', bg='1F3864')
    ws.merge_cells('R1:W1')

    # Satır 2: sütun başlıkları
    param_hdrs = ['q', 'T₀', 'α', 'ρ', 'Kombo #']
    result_hdrs = ['f_best\n(ort)', 'f_best\n(std)', 'f_best\n(min)',
                   'İyileştirme\n%', 'CPU\n(sn)', 'OnTime\n(ort)']
    for c, h in enumerate(param_hdrs, 1):
        sh(ws.cell(2, c), h, bg='2E75B6', sz=8, wrap=True)
    for grp_c in [6, 12, 18]:
        for c, h in enumerate(result_hdrs, grp_c):
            bg = '375623' if grp_c==6 else ('833C00' if grp_c==12 else '1F3864')
            sh(ws.cell(2, c), h, bg=bg, sz=8, wrap=True)

    # Sıralama için mean değerleri topla (renk kodlaması için)
    means_by_inst = {nm: [] for nm in TEST_INSTANCES}
    for combo_key, combo_data in all_data.items():
        for nm in TEST_INSTANCES:
            if nm in combo_data:
                means_by_inst[nm].append(combo_data[nm]['mean'])

    # Satırları yaz
    combos = list(all_data.keys())
    ri = 3
    for combo_idx, combo_key in enumerate(combos, 1):
        combo_data = all_data[combo_key]
        q, T0, alpha, rho = combo_key
        bg_row = alt(ri)

        # Parametre sütunları
        for c, v in enumerate([q, T0, alpha, rho, combo_idx], 1):
            sd(ws.cell(ri, c), v, bold=(c<=4), bg=bg_row)

        # Sonuç sütunları (her instance için)
        for grp_c, nm in zip([6, 12, 18], TEST_INSTANCES):
            if nm not in combo_data:
                for c in range(grp_c, grp_c + 6):
                    sd(ws.cell(ri, c), 'N/A', bg=bg_row)
                continue
            res = combo_data[nm]
            vals = [res['mean'], res['std'], res['best'],
                    res['impr_pct'], res['mean_cpu'], res['mean_ot']]
            fmts = ['0.000', '0.000', '0.000', '0.0', '0.00', '0.0']

            # mean değeri için renk sıralaması
            sorted_means = sorted(means_by_inst[nm])
            rank = sorted_means.index(res['mean']) if res['mean'] in sorted_means else 0
            bg_c, fg_c = rank_color(rank, len(sorted_means))

            for c, (v, fmt) in enumerate(zip(vals, fmts), grp_c):
                use_bg = bg_c if c == grp_c else bg_row   # sadece mean sütunu renkli
                use_fg = fg_c if c == grp_c else '000000'
                sd(ws.cell(ri, c), v, fmt=fmt, color=use_fg, bg=use_bg,
                   bold=(c == grp_c))
        ri += 1

    autofit(ws)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 6

# ════════════════════════════════════════════════════════════
#  SHEET 2–4: Her Instance için Özet (Tek Parametre Etkisi)
# ════════════════════════════════════════════════════════════
def sheet_single_param(wb, all_data, param_name, param_vals, inst_name,
                       sheet_name, other_defaults):
    """
    Tek parametrenin etkisini gösterir; diğerleri varsayılan değerde tutulur.
    """
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = 'A2'

    title = (f'{param_name} Parametresi Etkisi — {inst_name}  '
             f'(diğer sabitler: '
             + ', '.join(f'{k}={v}' for k, v in other_defaults.items()) + ')')
    sh(ws.cell(1, 1), title, bg='1F3864', sz=10)
    ws.merge_cells(f'A1:{get_column_letter(2 + N_SEEDS + 3)}1')

    hdrs = [param_name] + [f'Seed {i}' for i in range(N_SEEDS)] + \
           ['Ortalama', 'Std Sapma', 'En İyi', 'İyileştirme%', 'Tavsiye']
    for c, h in enumerate(hdrs, 1): sh(ws.cell(2, c), h, sz=9)

    best_mean = float('inf'); best_pv = None
    rows_data = []
    for pv in param_vals:
        key = tuple(other_defaults[p] if p != param_name else pv
                    for p in ['q', 'T0', 'alpha', 'rho'])
        if key not in all_data or inst_name not in all_data[key]:
            rows_data.append((pv, None))
            continue
        res = all_data[key][inst_name]
        seed_vals = [r['f_best'] for r in res['runs']]
        rows_data.append((pv, res, seed_vals))
        if res['mean'] < best_mean:
            best_mean = res['mean']; best_pv = pv

    for ri, item in enumerate(rows_data, 3):
        pv = item[0]; fill = alt(ri)
        if item[1] is None:
            sd(ws.cell(ri, 1), pv, bold=True, bg=fill)
            for c in range(2, len(hdrs)+1): sd(ws.cell(ri, c), 'N/A', bg=fill)
            continue
        res, seed_vals = item[1], item[2]
        is_best = (pv == best_pv)
        row_bg = 'C6EFCE' if is_best else fill
        sd(ws.cell(ri, 1), pv, bold=True, bg=row_bg)
        for c, v in enumerate(seed_vals, 2):
            sd(ws.cell(ri, c), v, fmt='0.000', bg=row_bg)
        extras = [res['mean'], res['std'], res['best'], res['impr_pct'],
                  f'← Tavsiye (ort={res["mean"]:.3f})' if is_best else '']
        for c, v in enumerate(extras, 2 + N_SEEDS):
            clr = '375623' if is_best and c == 2+N_SEEDS+4 else '000000'
            sd(ws.cell(ri, c), v,
               fmt='0.000' if isinstance(v, float) and c < 2+N_SEEDS+4 else None,
               color=clr, bold=is_best, bg=row_bg)

    note_row = 3 + len(rows_data) + 1
    ws.cell(note_row, 1).value = (
        f'Not: Her değer {N_SEEDS} bağımsız çalıştırmanın ortalamasıdır. '
        f'n_iter={N_ITER}, diğer parametreler sabit.')
    ws.cell(note_row, 1).font = Font(name='Calibri', italic=True, size=8)
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=len(hdrs))
    autofit(ws)

# ════════════════════════════════════════════════════════════
#  SHEET 5: En İyi 10 Kombinasyon (instance başına)
# ════════════════════════════════════════════════════════════
def sheet_top10(wb, all_data):
    ws = wb.create_sheet('En_Iyi_10_Kombinasyon')
    ws.freeze_panes = 'A3'
    sh(ws.cell(1, 1), 'Her Instance İçin En İyi 10 Parametre Kombinasyonu',
       bg='1F3864', sz=11)
    ws.merge_cells('A1:M1')

    hdrs = ['Sıra','Instance','q','T₀','α','ρ',
            'f_best (ort)','f_best (std)','f_best (min)','İyileştirme%',
            'CPU (sn)','OnTime','Notlar']
    for c, h in enumerate(hdrs, 1): sh(ws.cell(2, c), h, sz=8, wrap=True)
    ws.row_dimensions[2].height = 28

    ri = 3
    for nm in TEST_INSTANCES:
        # Bu instance için tüm sonuçları sırala
        inst_results = []
        for combo_key, combo_data in all_data.items():
            if nm not in combo_data: continue
            q, T0, alpha, rho = combo_key
            res = combo_data[nm]
            inst_results.append((res['mean'], combo_key, res))
        inst_results.sort(key=lambda x: x[0])

        for rank, (mean_val, combo_key, res) in enumerate(inst_results[:10], 1):
            q, T0, alpha, rho = combo_key
            fill = alt(ri)
            bg_c, fg_c = rank_color(rank-1, 10)
            note = '★ Tavsiye Edilen' if rank == 1 else (
                   '✓ İyi Alternatif' if rank <= 3 else '')
            row = [rank, nm, q, T0, alpha, rho,
                   res['mean'], res['std'], res['best'],
                   res['impr_pct'], res['mean_cpu'], res['mean_ot'], note]
            for c, v in enumerate(row, 1):
                use_bg = bg_c if rank <= 3 else fill
                use_fg = fg_c if rank <= 3 else ('375623' if '★' in str(v) else '000000')
                sd(ws.cell(ri, c), v,
                   fmt='0.000' if isinstance(v,float) and c in [7,8,9] else
                       ('0.0' if isinstance(v,float) and c in [10,12] else
                        ('0.00' if isinstance(v,float) and c==11 else None)),
                   color=use_fg, bold=(rank==1), bg=use_bg)
            ri += 1
        # Grup ayracı
        for c in range(1, len(hdrs)+1):
            ws.cell(ri, c).fill = _fill('2E75B6')
            ws.cell(ri, c).border = _border()
        ri += 1

    autofit(ws)

# ════════════════════════════════════════════════════════════
#  SHEET 6: İkili Etkileşim (q × alpha heat map tarzı)
# ════════════════════════════════════════════════════════════
def sheet_interaction(wb, all_data, inst_name):
    ws = wb.create_sheet(f'q_alpha_Etkilesim_{inst_name}')
    sh(ws.cell(1, 1),
       f'q × α Etkileşim Tablosu — {inst_name}  (T₀=6.0, ρ=0.20 sabit)',
       bg='1F3864', sz=10)
    ws.merge_cells(f'A1:{get_column_letter(1+len(GRID["alpha"])+1)}1')

    # Başlıklar
    sh(ws.cell(2, 1), 'q \\ α', bg='2E75B6', sz=9)
    for c, a in enumerate(GRID['alpha'], 2):
        sh(ws.cell(2, c), str(a), bg='2E75B6', sz=9)
    sh(ws.cell(2, len(GRID['alpha'])+2), 'En iyi α', bg='833C00', sz=9)

    for ri, q in enumerate(GRID['q'], 3):
        sh(ws.cell(ri, 1), str(q), bg='BDD7EE', fg='000000', sz=9)
        row_vals = []
        for c, alpha in enumerate(GRID['alpha'], 2):
            # T0=6.0, rho=0.20 sabit
            key = (q, 6.0, alpha, 0.20)
            if key in all_data and inst_name in all_data[key]:
                v = all_data[key][inst_name]['mean']
                row_vals.append((v, alpha))
            else:
                v = None; row_vals.append((None, alpha))
            if v is not None:
                # Tüm alpha değerleri içinde relatif sıra
                all_v = []
                for aa in GRID['alpha']:
                    k2 = (q, 6.0, aa, 0.20)
                    if k2 in all_data and inst_name in all_data[k2]:
                        all_v.append(all_data[k2][inst_name]['mean'])
                if all_v:
                    rank = sorted(all_v).index(v)
                    bg_c, fg_c = rank_color(rank, len(all_v))
                    sd(ws.cell(ri, c), v, fmt='0.000',
                       color=fg_c, bg=bg_c, bold=(rank==0))
            else:
                sd(ws.cell(ri, c), 'N/A', bg='F2F2F2')

        # En iyi alpha bu q için
        valid = [(v, a) for v, a in row_vals if v is not None]
        if valid:
            best_v, best_a = min(valid, key=lambda x: x[0])
            sd(ws.cell(ri, len(GRID['alpha'])+2), f'α={best_a} ({best_v:.3f})',
               color='375623', bold=True, bg='C6EFCE')

    # Not
    nr = 3 + len(GRID['q']) + 1
    ws.cell(nr, 1).value = 'Renk: Yeşil = düşük gecikme (iyi) | Kırmızı = yüksek gecikme (kötü)'
    ws.cell(nr, 1).font = Font(name='Calibri', italic=True, size=8, color='595959')
    ws.merge_cells(start_row=nr, start_column=1, end_row=nr,
                   end_column=len(GRID['alpha'])+2)
    autofit(ws)

# ════════════════════════════════════════════════════════════
#  SHEET 7: Tavsiye Özeti
# ════════════════════════════════════════════════════════════
def sheet_recommendation(wb, all_data):
    ws = wb.create_sheet('Parametre_Tavsiyesi')
    sh(ws.cell(1,1), 'ARP-SA Parametre Tavsiyesi — Özet', bg='1F3864', sz=12)
    ws.merge_cells('A1:G1')
    ws.row_dimensions[1].height = 28

    for c, h in enumerate(['Instance','Tavsiye q','Tavsiye T₀',
                            'Tavsiye α','Tavsiye ρ',
                            'Beklenen f_best','Beklenen İyileştirme%'], 1):
        sh(ws.cell(2, c), h, sz=9)

    ri = 3
    for nm in TEST_INSTANCES:
        # En iyi kombinasyonu bul
        best = min(
            ((res['mean'], combo_key, res)
             for combo_key, cd in all_data.items()
             if nm in cd
             for res in [cd[nm]]),
            key=lambda x: x[0], default=None)
        if best is None: continue
        mean_val, (q,T0,alpha,rho), res = best
        fill = 'C6EFCE'
        row = [nm, q, T0, alpha, rho,
               round(mean_val, 3), res['impr_pct']]
        for c, v in enumerate(row, 1):
            sd(ws.cell(ri, c), v,
               fmt='0.000' if isinstance(v,float) and c==6 else
                   ('0.0' if isinstance(v,float) and c==7 else None),
               bold=True, color='375623', bg=fill)
        ri += 1

    # Genel tavsiye notu
    ri += 1
    notes = [
        'q parametresi: Instance boyutuna göre ölçeklenir (önerilen: n//4).',
        'T₀ (başlangıç sıcaklığı): Yüksek değer erken iterasyonlarda daha fazla keşif sağlar.',
        'α (soğutma hızı): 0.97 dengeli; 0.99 daha yavaş soğur (daha uzun keşif).',
        'ρ (ağırlık güncelleme): Düşük değer geçmişe daha fazla ağırlık verir.',
    ]
    for note in notes:
        ws.cell(ri, 1).value = '• ' + note
        ws.cell(ri, 1).font = Font(name='Calibri', size=9, color='404040')
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=7)
        ri += 1

    autofit(ws)

# ════════════════════════════════════════════════════════════
#  ANA FONKSİYON
# ════════════════════════════════════════════════════════════
def main():
    print('=' * 68)
    print('  ARP-SA Parametre Optimizasyonu — Grid Search')
    print('=' * 68)

    # Instance'ları yükle
    instances = {}
    for nm in TEST_INSTANCES:
        fp = os.path.join(INST_DIR, f'{nm}.json')
        if not os.path.exists(fp):
            print(f'  UYARI: {nm}.json bulunamadı, atlanıyor.')
            continue
        with open(fp) as f:
            instances[nm] = json.load(f)
    if not instances:
        print('  HATA: Hiç instance bulunamadı. Önce code1_instances.py çalıştırın.')
        return

    # Kombinasyon sayısı
    combos = list(itertools.product(
        GRID['q'], GRID['T0'], GRID['alpha'], GRID['rho']))
    total = len(combos) * len(instances) * N_SEEDS
    print(f'\n  Toplam kombinasyon : {len(combos)}')
    print(f'  Test instance sayısı: {len(instances)}')
    print(f'  Seed sayısı        : {N_SEEDS}')
    print(f'  Toplam ALNS çalışması: {total}')
    print(f'  n_iter (sabit)     : {N_ITER}')
    print()

    # Grid search
    all_data = {}  # {(q,T0,alpha,rho): {inst_name: result_dict}}
    done = 0
    start_total = time.time()

    for ci, (q, T0, alpha, rho) in enumerate(combos, 1):
        combo_key = (q, T0, alpha, rho)
        all_data[combo_key] = {}
        for nm, inst in instances.items():
            res = run_combo(inst, q, T0, alpha, rho)
            all_data[combo_key][nm] = res
            done += N_SEEDS
        elapsed = time.time() - start_total
        remaining = elapsed / done * (total - done) if done > 0 else 0
        print(f'  Kombo {ci:>3}/{len(combos)}  '
              f'q={q} T0={T0:.1f} α={alpha} ρ={rho}  '
              + '  '.join(f'{nm}:{all_data[combo_key][nm]["mean"]:7.3f}'
                          for nm in instances)
              + f'  [kalan ~{remaining/60:.1f}dk]')

    print(f'\n  Toplam süre: {(time.time()-start_total)/60:.1f} dakika')

    # Varsayılan parametre değerleri (tek parametre analizi için)
    defaults = {'q': 4, 'T0': 6.0, 'alpha': 0.97, 'rho': 0.20}

    # Excel oluştur
    print('\n  Excel dosyası oluşturuluyor...')
    wb = Workbook(); wb.remove(wb.active)

    sheet_all_combos(wb, all_data, instances)

    # Her parametre için tek-parametre analizi (M01 üzerinde)
    main_inst = 'M01' if 'M01' in instances else list(instances.keys())[0]
    for pname, pvals in GRID.items():
        others = {k: defaults[k] for k in GRID if k != pname}
        sheet_name = f'{pname}_Etkisi'
        sheet_single_param(wb, all_data, pname, pvals, main_inst,
                           sheet_name, defaults)

    sheet_top10(wb, all_data)
    sheet_interaction(wb, all_data, main_inst)
    sheet_recommendation(wb, all_data)

    out = os.path.join(BASE, 'ARP_SA_ParamTuning.xlsx')
    wb.save(out)

    print(f'  Kaydedildi → {out}')
    print('  Sayfalar  : ' + ', '.join(wb.sheetnames))
    return all_data

if __name__ == '__main__':
    main()