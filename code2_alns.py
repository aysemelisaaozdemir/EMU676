#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jun  9 18:57:08 2026

@author: melisaozdemir
"""

#!/usr/bin/env python3
"""
Kod 2 — ARP-SA ALNS Çalıştırıcı
Kod 1'in ürettiği instance JSON'larını okur, ALNS çalıştırır,
sonuçları ARP_SA_Results.xlsx'e kaydeder.

Gereksinim: arp_sa_v2.py ile aynı klasörde olmalı.
"""
import os, json, time, math, statistics, sys
import arp_sa_v2 as _m
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
INST_DIR = os.path.join(BASE, 'instances')

# ── Excel stil yardımcıları ──────────────────────────────────
def _fill(hex_): return PatternFill('solid', start_color=hex_)
def _border():
    t = Side(style='thin', color='BFBFBF')
    return Border(left=t, right=t, top=t, bottom=t)
def set_hdr(cell, txt, fg='FFFFFF', bg='1F3864', sz=9, bold=True, wrap=False):
    cell.value = txt
    cell.font = Font(name='Calibri', bold=bold, color=fg, size=sz)
    cell.fill = _fill(bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    cell.border = _border()
def set_data(cell, val, fmt=None, bold=False, align='center', color='000000'):
    cell.value = val
    cell.font = Font(name='Calibri', bold=bold, size=9, color=color)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = _border()
    if fmt: cell.number_format = fmt
def alt(ri): return _fill('DCE6F1') if ri % 2 == 0 else _fill('FFFFFF')
def autofit(ws, extra=2):
    for col in ws.columns:
        mx = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = mx + extra

# ── Instance yükleme & ALNS global patching ──────────────────
def _patch(inst):
    """arp_sa_v2 global değişkenlerini geçici instance verisine bağla."""
    orig = {k: getattr(_m, k) for k in ['ORDERS','DUE','PROC','FIX','SU','NVEH']}
    _m.ORDERS = inst['orders']
    _m.DUE    = {p: float(d) for p, d in inst['due'].items()}
    _m.PROC   = inst['proc']
    _m.FIX    = inst['fixtures']
    _m.SU     = inst['setup']
    _m.NVEH   = inst['n_vehicles']
    return orig

def _restore(orig):
    for k, v in orig.items(): setattr(_m, k, v)

def run_instance(inst, n_iter, n_seeds=5):
    q = max(3, inst['n_orders'] // 4)
    runs = []
    for seed in range(n_seeds):
        orig = _patch(inst)
        try:
            t0 = time.time()
            a, s, r, fb, fi, hist, op = _m.alns(
                n_iter=n_iter, q=q, T0=6., alpha=.97, seed=seed, verbose=False)
            cpu = time.time() - t0
            _, r_arr, C, T = _m.evaluate(a, s, r)
        finally:
            _restore(orig)
        runs.append({
            'seed': seed, 'f_init': round(fi, 3), 'f_best': round(fb, 3),
            'impr_pct': round(100*(fi-fb)/max(fi,1e-9), 1),
            'cpu_s': round(cpu, 3),
            'on_time': sum(1 for p in inst['orders'] if T[p] < 1e-6),
            'tardiness': {p: round(T[p], 2) for p in inst['orders']},
            'completion': {p: round(C[p], 2) for p in inst['orders']},
            'assignments': {p: dict(a[p]) for p in inst['orders']},
            'routes': r,
            'convergence': hist,
            'op_stats': op,
        })
    return runs

# ── Sheet 1: Özet İstatistikler ──────────────────────────────
def sheet_summary(wb, all_results):
    ws = wb.create_sheet('Ozet_Istatistikler'); ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 24
    hdrs = ['Instance','|P|','Zorluk','f_init','f_best (ort)',
            'f_best (std)','f_best (min)','f_best (maks)',
            'İyileştirme %','OnTime (ort)','CPU ort (sn)']
    for c, h in enumerate(hdrs, 1): set_hdr(ws.cell(1, c), h, wrap=True, sz=8)
    for ri, (nm, data) in enumerate(all_results.items(), 2):
        fb_vals = [r['f_best'] for r in data['runs']]
        on_vals = [r['on_time'] for r in data['runs']]
        cp_vals = [r['cpu_s']  for r in data['runs']]
        impr    = round(100*(data['runs'][0]['f_init']-min(fb_vals))/max(data['runs'][0]['f_init'],1e-9),1)
        row = [nm, data['n_orders'], data['difficulty'],
               data['runs'][0]['f_init'],
               round(statistics.mean(fb_vals),3),
               round(statistics.stdev(fb_vals) if len(fb_vals)>1 else 0.,3),
               round(min(fb_vals),3), round(max(fb_vals),3),
               impr,
               round(statistics.mean(on_vals),1),
               round(statistics.mean(cp_vals),2)]
        fill = alt(ri)
        for c, v in enumerate(row, 1):
            fmt = '0.000' if isinstance(v, float) and c in [4,5,6,7,8] else (
                  '0.0%' if c==9 else None)
            col_c = 'C00000' if c==9 and impr<30 else ('70AD47' if c==9 and impr>80 else '000000')
            set_data(ws.cell(ri, c), v, fmt=fmt, color=col_c,
                     bold=(c==1 or c==9))
            ws.cell(ri, c).fill = fill
        # Grup ayracı (boyut değişiminde ince çizgi)
    autofit(ws)

# ── Sheet 2: Detaylı Çalıştırma Sonuçları ────────────────────
def sheet_detailed(wb, all_results):
    ws = wb.create_sheet('Detayli_Calistirmalar'); ws.freeze_panes = 'C2'
    # Sipariş sütunları için başlıklar (dinamik — en büyük n)
    max_n = max(d['n_orders'] for d in all_results.values())
    order_cols = [f'P{i+1:02d}' for i in range(max_n)]
    hdrs_fixed = ['Instance','|P|','Seed','f_init','f_best','İyileştirme%','CPU (sn)','OnTime']
    for c, h in enumerate(hdrs_fixed, 1): set_hdr(ws.cell(1, c), h, sz=8)
    c0 = len(hdrs_fixed) + 1
    set_hdr(ws.cell(1, c0), 'Sipariş Bazında Gecikme T_p (saat)',
            bg='2E75B6', sz=8)
    ws.merge_cells(start_row=1, start_column=c0, end_row=1,
                   end_column=c0+max_n-1)
    ws2_row = 2
    # Satır 2: sipariş başlıkları
    for c, h in enumerate(hdrs_fixed, 1):
        set_hdr(ws.cell(2, c), '', bg='F2F2F2', fg='000000', sz=8)
    for c, p in enumerate(order_cols, c0):
        set_hdr(ws.cell(2, c), p, bg='BDD7EE', fg='000000', sz=8)
    ws.freeze_panes = 'A3'
    ri = 3
    for nm, data in all_results.items():
        for run in data['runs']:
            fill = alt(ri)
            base = [nm, data['n_orders'], run['seed'],
                    run['f_init'], run['f_best'], run['impr_pct'],
                    run['cpu_s'], run['on_time']]
            for c, v in enumerate(base, 1):
                set_data(ws.cell(ri, c), v, fmt='0.000' if isinstance(v,float) and c in [4,5] else None)
                ws.cell(ri, c).fill = fill
            for c, p in enumerate(order_cols, c0):
                v = run['tardiness'].get(p, '')
                fmt = '0.00' if isinstance(v,(int,float)) else None
                clr = 'C00000' if isinstance(v,(int,float)) and v>0 else '000000'
                set_data(ws.cell(ri, c), v, fmt=fmt, color=clr)
                ws.cell(ri, c).fill = fill
            ri += 1
    autofit(ws)

# ── Sheet 3: En İyi Çözüm Atamaları ─────────────────────────
def sheet_best_solutions(wb, all_results):
    ws = wb.create_sheet('En_Iyi_Atamalar'); ws.freeze_panes = 'B2'
    hdrs = ['Instance','Sipariş','Due (saat)','C_p (saat)','T_p (saat)',
            'M Ataması','E Ataması','W Ataması','Gecikme Durumu']
    for c, h in enumerate(hdrs, 1): set_hdr(ws.cell(1, c), h, sz=8)
    ri = 2
    for nm, data in all_results.items():
        # En iyi seed'i bul
        best_run = min(data['runs'], key=lambda x: x['f_best'])
        inst_due  = data['due']
        for p in data['orders']:
            fill = alt(ri)
            T_p  = best_run['tardiness'].get(p, 0.)
            C_p  = best_run['completion'].get(p, 0.)
            asgn = best_run['assignments'].get(p, {})
            status = '✓ Zamanında' if T_p < 1e-6 else f'✗ {T_p:.2f}h gecikme'
            row = [nm, p, inst_due.get(p,''), round(C_p,2), round(T_p,2),
                   asgn.get('M',''), asgn.get('E',''), asgn.get('W',''), status]
            for c, v in enumerate(row, 1):
                clr = 'C00000' if '✗' in str(v) else ('375623' if '✓' in str(v) else '000000')
                set_data(ws.cell(ri,c), v, color=clr, bold=(c==1),
                         fmt='0.00' if isinstance(v,float) else None)
                ws.cell(ri,c).fill = fill
            ri += 1
    autofit(ws)

# ── Sheet 4: Operatör Analizi ─────────────────────────────────
def sheet_operators(wb, all_results):
    ws = wb.create_sheet('Operator_Analizi'); ws.freeze_panes = 'C2'
    # Satır 1: üst başlıklar
    set_hdr(ws.cell(1,1),'Instance'); set_hdr(ws.cell(1,2),'Seed')
    d_names=['Random','WorstTard','Bottleneck']
    r_names=['Greedy','Regret']
    c=3
    for grp,names in [('Yok Etme Operatörleri (Çağrı/Ağırlık/İyileştirme)',d_names),
                       ('Onarma Operatörleri (Çağrı/Ağırlık/İyileştirme)',r_names)]:
        ws.merge_cells(start_row=1,start_column=c,end_row=1,end_column=c+len(names)*3-1)
        set_hdr(ws.cell(1,c),grp,bg='2E75B6',sz=8)
        for nm in names:
            set_hdr(ws.cell(2,c),f'{nm}\nÇağrı',bg='BDD7EE',fg='000000',sz=8,wrap=True)
            set_hdr(ws.cell(2,c+1),f'{nm}\nAğırlık',bg='BDD7EE',fg='000000',sz=8,wrap=True)
            set_hdr(ws.cell(2,c+2),f'{nm}\nİyileşt.',bg='BDD7EE',fg='000000',sz=8,wrap=True)
            c+=3
    set_hdr(ws.cell(2,1),'',bg='F2F2F2',fg='000000')
    set_hdr(ws.cell(2,2),'',bg='F2F2F2',fg='000000')
    ws.freeze_panes='A3'
    ri=3
    for nm, data in all_results.items():
        for run in data['runs']:
            fill=alt(ri)
            set_data(ws.cell(ri,1),nm,bold=True); ws.cell(ri,1).fill=fill
            set_data(ws.cell(ri,2),run['seed']);  ws.cell(ri,2).fill=fill
            op=run['op_stats']
            c=3
            for i,_ in enumerate(d_names):
                set_data(ws.cell(ri,c),op['d_calls'][i]);          ws.cell(ri,c).fill=fill
                set_data(ws.cell(ri,c+1),round(op['d_weights'][i],3),'0.000'); ws.cell(ri,c+1).fill=fill
                set_data(ws.cell(ri,c+2),round(op['d_impr'][i],2),'0.00');     ws.cell(ri,c+2).fill=fill
                c+=3
            for i,_ in enumerate(r_names):
                set_data(ws.cell(ri,c),op['r_calls'][i]);          ws.cell(ri,c).fill=fill
                set_data(ws.cell(ri,c+1),round(op['r_weights'][i],3),'0.000'); ws.cell(ri,c+1).fill=fill
                set_data(ws.cell(ri,c+2),round(op['r_impr'][i],2),'0.00');     ws.cell(ri,c+2).fill=fill
                c+=3
            ri+=1
    autofit(ws)

# ── Sheet 5: Parametre Analizi (q) ───────────────────────────
def sheet_param(wb):
    ws = wb.create_sheet('Parametre_Analizi_q')
    ws.row_dimensions[1].height=20
    set_hdr(ws.cell(1,1),'q Parametresi Duyarlılık Analizi — M01 Instance (5 seed × 600 iter)',
            bg='1F3864',sz=10)
    ws.merge_cells('A1:H1')
    hdrs2=['q Değeri','Seed 0','Seed 1','Seed 2','Seed 3','Seed 4','Ortalama','Standart Sapma']
    for c,h in enumerate(hdrs2,1): set_hdr(ws.cell(2,c),h,sz=9)
    # M01 için q değerlerini çalıştır
    m01_path=os.path.join(INST_DIR,'M01.json')
    with open(m01_path) as f: m01=json.load(f)
    ri=3
    for q in [2,3,4,6,9]:
        vals=[]
        for seed in range(5):
            orig=_patch(m01)
            try:
                _,_,_,fb,_,_,_=_m.alns(n_iter=600,q=q,seed=seed,verbose=False)
            finally: _restore(orig)
            vals.append(round(fb,3))
        fill=alt(ri)
        row=[q]+vals+[round(statistics.mean(vals),3),
                      round(statistics.stdev(vals) if len(vals)>1 else 0.,3)]
        for c,v in enumerate(row,1):
            set_data(ws.cell(ri,c),v,'0.000' if isinstance(v,float) else None,
                     bold=(c==1)); ws.cell(ri,c).fill=fill
        ri+=1
    # Notlar
    ws.cell(ri+1,1).value='Not: q = yok etme operatörü başına kaldırılan (p,τ) çifti sayısı.'
    ws.cell(ri+1,1).font=Font(name='Calibri',italic=True,size=9)
    autofit(ws)

# ── Sheet 6: Yakınsama ───────────────────────────────────────
def sheet_convergence(wb, all_results):
    ws = wb.create_sheet('Yakinasama_Verileri')
    set_hdr(ws.cell(1,1),'Instance'); set_hdr(ws.cell(1,2),'Seed')
    set_hdr(ws.cell(1,3),'f_init')
    # Iterasyon sütunları (maksimum uzunluk)
    max_hist=max(len(r['convergence']) for d in all_results.values() for r in d['runs'])
    step=max(1,max_hist//50)  # her 'step' iterasyonu bir sütun
    iter_cols=list(range(0,max_hist,step))
    for c,it in enumerate(iter_cols,4): set_hdr(ws.cell(1,c),f'Iter {it}',sz=8)
    ri=2
    for nm,data in all_results.items():
        for run in data['runs']:
            fill=alt(ri)
            set_data(ws.cell(ri,1),nm,bold=True); ws.cell(ri,1).fill=fill
            set_data(ws.cell(ri,2),run['seed']);  ws.cell(ri,2).fill=fill
            set_data(ws.cell(ri,3),run['f_init'],'0.000'); ws.cell(ri,3).fill=fill
            hist=run['convergence']
            for c,it in enumerate(iter_cols,4):
                v=hist[it] if it<len(hist) else hist[-1]
                set_data(ws.cell(ri,c),round(v,3),'0.000')
                ws.cell(ri,c).fill=fill
            ri+=1
    autofit(ws)

# ── Ana fonksiyon ─────────────────────────────────────────────
def main():
    print("="*65)
    print("  ARP-SA — ALNS Çalıştırıcı (Kod 2)")
    print("="*65)

    # Instance JSON'larını yükle
    json_files=sorted([f for f in os.listdir(INST_DIR) if f.endswith('.json')])
    inst_order=['M01','S01','S02','S03','S04','S05','S06','S07','S08',
                'M02','M03','M04','M05','M06','M07','M08',
                'L01','L02','L03','L04','L05','L06']
    # Sadece mevcut dosyaları al
    available=[nm for nm in inst_order
               if os.path.exists(os.path.join(INST_DIR,f'{nm}.json'))]

    all_results={}
    print(f"\n  {'Instance':<8} {'|P|':>4}  {'f_init':>8}  {'f_best':>8}  {'İmpr%':>7}  {'CPU':>7}  {'OnTime':>8}")
    print("  "+"-"*60)

    for nm in available:
        with open(os.path.join(INST_DIR,f'{nm}.json')) as f:
            inst=json.load(f)
        n=inst['n_orders']
        n_iter=400 if n<=5 else (500 if n<=7 else (600 if n<=10 else (800 if n<=15 else 1000)))
        runs=run_instance(inst,n_iter=n_iter,n_seeds=5)
        fb_vals=[r['f_best'] for r in runs]
        all_results[nm]={
            'n_orders':n,'difficulty':inst.get('difficulty','?'),
            'due':inst['due'],'orders':inst['orders'],
            'runs':runs,
        }
        impr=runs[0]['impr_pct']
        print(f"  {nm:<8} {n:>4}  {runs[0]['f_init']:>8.2f}  "
              f"{statistics.mean(fb_vals):>8.3f}  {impr:>7.1f}%  "
              f"{statistics.mean(r['cpu_s'] for r in runs):>6.2f}s  "
              f"{statistics.mean(r['on_time'] for r in runs):>5.1f}/{n}")

    # Excel oluştur
    print("\n  Excel dosyası oluşturuluyor...")
    wb=Workbook(); wb.remove(wb.active)
    sheet_summary(wb,all_results)
    sheet_detailed(wb,all_results)
    sheet_best_solutions(wb,all_results)
    sheet_operators(wb,all_results)
    print("  Parametre analizi (q) çalışıyor...")
    sheet_param(wb)
    sheet_convergence(wb,all_results)

    out=os.path.join(BASE,'ARP_SA_Results.xlsx')
    wb.save(out)
    print(f"  Excel → {out}")
    print("\n  Sayfalar:")
    for s in wb.sheetnames: print(f"    · {s}")

if __name__=='__main__':
    main()